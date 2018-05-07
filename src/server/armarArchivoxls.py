import openpyxl
import cx_Oracle as ORA

US = 'owner_rafam'
PASS = 'ownerdba'
HOST = '172.22.20.151:1521'
DB = 'MBROWN'


def pedirDatos():
    orden = "Select o.docket_id,dep.nombre, SUM(o.amount), SUM(o.amount*(o.hour_value/(180*60))*o.mod),o.mod \
    from overtimes o, DEPENDENCIAS dep \
    where o.o_date >= TO_DATE( '2018-04-01', 'YYYY-MM-DD' ) \
    and o.o_date <= TO_DATE( '2018-04-30', 'YYYY-MM-DD' ) \
    and o.dependence_id = dep.codigo and dep.JURISDICCION = '{}' \
    group by o.docket_id,o.mod"

    con,cur = conectar()

    secretarias = ['1110119000','1110108000','1110122000','11101200000']
    nombres = ['Gobierno','Salud','Infraestructura','Seguridad']
    
    datos = []

    for secretaria in range(len(secretarias)):
        
        cur.execute(orden.format(secretarias[secretaria]).encode('utf8'))

        dic = {}

        for res in cur.fetchall():
            if not (res[1] in dic):
                dic[res[1]] = [res[1],{}]
            
            if not (res[0] in dic[ res[1] ] [1]):
                dic[ res[1] ] [ 1 ] [ res[0] ] = [ res[0], [ [0,0],[0,0],[0,0],[0,0] ] ]
            


            if(res[4] == 1.5):
                dic[res[1]][1][res[0]][1][0] = [res[2],res[3]]
            elif(res[4] == 2):
                dic[res[1]][1][res[0]][1][1] = [res[2],res[3]]
            elif(res[4] == 12):
                dic[res[1]][1][res[0]][1][2] = [res[2],res[3]]
            elif(res[4] == 16):
                dic[res[1]][1][res[0]][1][3] = [res[2],res[3]]

        datos.append([nombres[secretaria],dic])

    return datos

#155,156,157, llamar a compras mariano. 0800 no anda


def armarArchivo(datos):

    archivo = openpyxl.Workbook()
    
    for secretaria in range(len(datos)):

        pagina = archivo.create_sheet(datos[secretaria][0])

        dependencias = datos[secretaria][1]

        fila = 1 #para controlar la altura en el archivo

        columnas = ['Legajo','Horas al 50%','Gasto al 50%','Horas al 100%'
        ,'Gasto al 100%','Horas al 50%N','Gasto al 50%N','Horas al 100%N','Gasto al 100%N']
        for x in range(len(columnas)):
            pagina.cell(row=fila,column=x + 1,value=columnas[x])
        
        fila += 2


        for dependencia in dependencias:

            pagina.cell(row=fila,column=1,value=dependencia[0])
            fila += 1

            for legajo in dependencia[1]:
                columna = 1

                pagina.cell(row=fila,column=columna,value=legajo[0])
                columna += 1
                for horas in legajo[1]:

                    pagina.cell(row= fila, column= columna, value = horas[0])
                    columna += 1
                    pagina.cell(row= fila, column= columna, value = horas[1])
                    columna += 1
                fila += 1
            
            fila += 2

    archivo.save('resumen abril.xlsx')



def main():
    datos = pedirDatos()
    armarArchivo(datos)


def conectar():
    con = ORA.connect(""+US + "/" + PASS + "@" + HOST + "/" + DB)
    cur = con.cursor()
    return con,cur



main()