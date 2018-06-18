#coding=latin
import Conectar as CON
import openpyxl
import pypyodbc as SQL



ordenSQL = 'SELECT \
c.[CHECKTIME] \
,c.[SENSORID], \
u.[NAME], \
u.[BADGENUMBER] \
FROM [rafam_relojes].[dbo].[CHECKINOUT] c \
inner join [rafam_relojes].[dbo].[USERINFO] u \
on c.USERID = u.USERID \
where c.[CHECKTIME] between "2018-05-01" and "2018-05-31" \
order by u.BADGENUMBER,c.CHECKTIME'

ordenORA = "SELECT j.denominacion, d.descripcion\
 from per_agentes_movimientos l \
 inner join dependencias d on l.dependencia_f = d.codigo \
 inner join jurisdicciones j on l.jurisdiccion = j.jurisdiccion \
 where l.legajo = {} and vigente = 'S'"

def conectar():
    con = SQL.connect('Driver={SQL Server};'
    'Server = 10.10.10.66;'
    'Database = rafam_relojes;'
    'uid=rafam_relojes;'
    'pwd=g9cwwS6w;')
    print 'creo el conector'
    cur = con.cursor()
    return con,cur



def pedir_horarios():
    con,cur = conectar()
    cur.execute(ordenSQL)
    fichadas = cur.fetchall()
    con.close()
    return fichadas

def pedir_datos_agente(legajo):
    con,cur = CON.conectar()
    
    cur.execute(ordenORA.format(legajo))
    info = cur.fetchall()[0]
    con.close()
    return info

def juntarDatos(horarios):

    datos = {} #estructura = {jurisdiccion:{dependencia:{legajo:[horarios]}}}
    for horario in horarios:
        legajo = horario[3]
        jurisdiccion,dependencia = pedir_datos_agente(legajo)
        
        if !(jurisdiccion in datos):
            datos[jurisdiccion] = {}
        if !(dependencia in datos[jurisdiccion]):
            datos[jurisdiccion][dependencia] = {}
        if !(legajo in datos[jurisdiccion][dependencia]):
            datos[jurisdiccion][dependencia][legajo] = []

        datos[jurisdiccion][dependencia][legajo].append([horario[0],horario[1],horario[2]])

    return datos

def crearArchivo():
    arch = openpyxl.Workbook()
    return arch

def cargar_datos(datos,arch):

    for jurisdiccion in datos:

        pagina = arch.create_sheet(jurisdiccion)

        altura = 1 
        for dependencia in datos[jurisdiccion]:

            dependencia = datos[jurisdiccion]
            pagina.cell(row=altura,column=1,value=dependencia)
            altura += 1
            for legajo in datos[jurisdiccion][dependencia]:

                pagina.cell(row=altura,column=1,value=legajo)
                altura += 1
                for horario in datos[jurisdiccion][dependencia][legajo]:
                    pagina.cell(row=altura,column=1,value=horario[0])
                    pagina.cell(row=altura,column=2,value=horario[1])
                    altura += 1
    
    arch.save('resumenFichadas.xlsx')

def main():
    arch = crearArchivo()
    horarios = pedir_horarios()
    datos = juntarDatos(horarios)
    cargar_datos(datos,arch)


main()



