from bottle import template, route, run, response, Bottle, hook, request,static_file,post,default_app
from json import dumps,loads
import datetime
import cx_Oracle as ORA
import openpyxl

US = 'owner_rafam'
PASS = 'ownerdba'
HOST = '172.22.20.151:1521'
DB = 'MBROWN'


@hook('after_request')
def enable_cors():
    """
    You need to add some headers to each request.
    Don't use the wildcard '*' for Access-Control-Allow-Origin in production.
    """
    response.add_header("Access-Control-Allow-Origin", "*")
    response.add_header("Access-Control-Allow-Methods", "POST,GET")
    response.add_header("Access-Control-Allow-Headers", "Origin, Accept, Content-Type")
    response.add_header("Access-Control-Max-Age", "1728000")
    

@post('/cors')
def lvambience():
    response.headers['Content-Type'] = 'application/json'
    return "[1]"


@route('/<:re:.*>', method='OPTIONS')
def dummy():
    return

def isoformatArgentino(string):
    fecha,hora = string.split('T')
    fecha = fecha[8:] + '/' + fecha[5:7] + '/' + fecha[0:4]
    return fecha + ' ' + hora

@route('/<modulo>')
def devolverModulo(modulo):
    return static_file(modulo,root="../../build/")

@route('/')
def devolverPagina():
    return static_file("index.html",root="../../build/")

def pedirPresentismo():

    lineaDato = "select a.docket_id,a.check_in,a.check_out, \
    (24 * (a.check_out - a.check_in) ), \
    concat(l.nombres,concat(' ',l.apellido), \
    j.denominacion,c.denominacion,h.valor \
    from attendance a \
    inner join legajos l on l.legajo = a.docket_id \
    inner join per_agentes_movimientos m on m.legajo = a.docket_id \
    inner join jurisdicciones j on m.jurisdiccion = j.jurisdiccion \
    inner join categorias c on  m.agrupamiento = c.agrupamiento \
    and c.categoria = m.categoria \
    and c.ejercicio = m.ejercicio \
    inner join modulo_horarios h on m.modulo_horario = h.codigo where c.jurisdiccion = '1110100000'\
    m.fecha_desde <= NOW() and (m.fecha_hasta > NOW() or m.fecha_hasta is null) " #now se puede cambiar por la fecha inicial, para asegurar que tenga mas sentido el modulo horario y la categoria
    
    if(secretariat != None):
        lineaDato += ' and m.jurisdiccion = %(secretariat)s'
    

    if(initial_date != None):
        lineaDato += ' and a.check_in >= %(initial_date)s and a.check_in <= %(end_date)s'

    if(category != None and group != None):
        lineaDato += ' and m.categoria = %(category)s and m.agrupamiento = %(group)s'


def armarArchivo():
    
    #recibe los mismos parametros que el de presentismo o le mando los datos desde la web.
    import openpyxl
    from openpyxl.utils import get_column_letter

    archivo = openpyxl.Workbook()
    hoja = archivo.active


    campos = ['Legajo','Ingreso','Egreso','Horas Trabajadas','Nombre','Secretaria','Categoria','Modulo Horario']
    
    tamCampos = [0,0,0,0,0,0,0,0]
    
    for x in range(len(campos)):
        
        hoja.cell(row=1, column=(x + 1), value=campos[x] )
        tamCampos[x] = len( campos[x] )

    for x in range( len(datos) ): #datos seria la lista de horarios a guardar.
        
        linea = datos[x]
        for y in range( len(campos) ):
            
            campo = linea[y]

            hoja.cell(row = (2 + x), column = ( y + 1), value=campo)
            if(tamCampos[y] < len(campo)):
                tamCampos[y] = len(campo)
    
    for x in range(len(tamCampos)):
        
        hoja.column_dimensions[ get_column_letter[x + 1] ].width = tamCampos[x]


    hoja.save('resumen.xlsx')

    return static_file('resumen.xlsx',root='./') #aca habria que poner le metodo de flask que mande archivos en blob    






def pedirDatos(desde,hasta):
    orden = "Select o.docket_id,dep.descripcion, SUM(o.amount), SUM(o.amount*(o.hour_value/(180*60))*o.mod),o.mod \
    from overtimes o, DEPENDENCIAS dep \
    where o.o_date >= TO_DATE( '{}', 'YYYY-MM-DD' ) \
    and o.o_date <= TO_DATE( '{}', 'YYYY-MM-DD' ) \
    and o.dependence_id = dep.codigo \
    and dep.jurisdiccion = '{}' \
    group by o.docket_id,dep.descripcion,o.mod"

    con,cur = conectar()

    ordenSecretarias = 'Select jurisdiccion,denominacion from jurisdicciones'
    cur.execute(ordenSecretarias)
    secretarias = cur.fetchall()
    
    
    datos = []

    for secretaria in range(len(secretarias)):
        
        cur.execute(orden.format(desde,hasta,secretarias[secretaria][0]).encode('utf8'))

        dic = {}

        for res in cur.fetchall():
            if not (res[1] in dic):
                dic[res[1]] = [res[1],{}]
            
            if not (res[0] in dic[ res[1] ] [1]):
                dic[ res[1] ] [ 1 ] [ res[0] ] = [ res[0], [ [0,0],[0,0],[0,0],[0,0] ] ]
            


            if(res[4] == 1.5):
                dic[res[1]][1][res[0]][1][0] = [res[2],res[3]]
            elif(res[4] == 2):
                dic[res[1]][1][res[0]][1][1] = [res[2],res[3]]
            elif(res[4] == 12):
                dic[res[1]][1][res[0]][1][2] = [res[2],res[3]]
            elif(res[4] == 16):
                dic[res[1]][1][res[0]][1][3] = [res[2],res[3]]

        datos.append([secretarias[secretaria][1],dic])

    return datos


def conectar():
    con = ORA.connect(""+US + "/" + PASS + "@" + HOST + "/" + DB)
    cur = con.cursor()
    return con,cur

def armarArchivo(datos):

    archivo = openpyxl.Workbook()
    
    for secretaria in range(len(datos)):

        pagina = archivo.create_sheet(datos[secretaria][0])

        dependencias = datos[secretaria][1]

        fila = 1 #para controlar la altura en el archivo

        columnas = ['Legajo','Horas al 50%','Gasto al 50%','Horas al 100%'
        ,'Gasto al 100%','Horas al 50%N','Gasto al 50%N','Horas al 100%N','Gasto al 100%N']
        for x in range(len(columnas)):
            pagina.cell(row=fila,column=x + 1,value=columnas[x])
        
        fila += 2


        for dependencia in dependencias.values():

            pagina.cell(row=fila,column=1,value=dependencia[0])
            fila += 1

            for legajo in dependencia[1].values():
                columna = 1

                pagina.cell(row=fila,column=columna,value=legajo[0])
                columna += 1
                for horas in legajo[1]:

                    pagina.cell(row= fila, column= columna, value = horas[0])
                    columna += 1
                    pagina.cell(row= fila, column= columna, value = horas[1])
                    columna += 1
                fila += 1
            
            fila += 2

    archivo.save('resumen.xlsx')


@route('/informe/excel',method='POST')
def main():
    datos = request.json
    desde = datos['initial_date']
    hasta = datos['end_date']
    info = pedirDatos(desde,hasta)
    armarArchivo(info)
    
    return static_file('resumen.xlsx',root='./')





run(host="0.0.0.0", port=3000)